import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Add src to path so app can be imported when run as a module (must be before app imports)
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

# noqa: E402 - path manipulation required before app imports
from app.database import Base, get_engine, get_session  # noqa: E402
from app.models import Activity, Settings, Task  # noqa: E402


def load_excel_data():
    import os

    home = str(Path.home())
    downloads = os.path.join(home, "Downloads/OneDrive_1_5-23-2026")
    project_root = ".."

    excel_files = []
    for year in [2022, 2023, 2024, 2025, 2026]:
        filename = f"DailyActivity_{year}.xlsx"
        # Check project root first, then Downloads
        project_path = os.path.join(project_root, filename)
        downloads_path = os.path.join(downloads, filename)

        if Path(project_path).exists():
            excel_files.append(project_path)
        elif Path(downloads_path).exists():
            excel_files.append(downloads_path)

    db = get_session()
    tasks_map = {}
    activity_count = 0

    print("Loading tasks from all files...")
    for excel_file in excel_files:
        if not Path(excel_file).exists():
            continue

        print(f"  Processing {Path(excel_file).name}...")
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        try:
            tasks_sheet = wb["task description"]
        except KeyError:
            continue

        for row in tasks_sheet.iter_rows(min_row=2, max_row=tasks_sheet.max_row, values_only=True):
            if row[0] is None:
                break
            task_name = row[0]
            task_type = row[1] if len(row) > 1 else None
            sub_type = row[2] if len(row) > 2 else None
            source = row[3] if len(row) > 3 else None
            links = row[6] if len(row) > 6 else None

            if not task_name or not isinstance(task_name, str):
                continue

            # Only add if task name doesn't already exist (keep first occurrence)
            existing = db.query(Task).filter(Task.name == task_name).first()
            if not existing:
                task = Task(
                    name=task_name,
                    type=task_type or "",
                    sub_type=sub_type or "",
                    source=source or "",
                    links=links,
                )
                db.add(task)
                db.flush()
                tasks_map[task_name] = task.id
            else:
                tasks_map[task_name] = existing.id

    db.commit()
    print(f"✓ Loaded {len(tasks_map)} unique tasks")

    print("Loading activities from all files...")
    for excel_file in excel_files:
        if not Path(excel_file).exists():
            continue

        print(f"  Processing {Path(excel_file).name}...")
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        try:
            activity_sheet = wb["activity"]
        except KeyError:
            continue

        file_activity_count = 0
        for row_idx, row in enumerate(
            activity_sheet.iter_rows(min_row=2, max_row=activity_sheet.max_row, values_only=True), 2
        ):
            date_cell = row[0]
            task_name_cell = row[3]
            start_time_cell = row[4]
            end_time_cell = row[5]
            comments_cell = row[7] if len(row) > 7 else None
            links_cell = row[8] if len(row) > 8 else None

            if date_cell is None or task_name_cell is None or start_time_cell is None:
                continue

            if not isinstance(task_name_cell, str) or not isinstance(date_cell, datetime):
                continue

            task_id = tasks_map.get(task_name_cell)
            if not task_id:
                continue

            date_val = date_cell.date() if isinstance(date_cell, datetime) else None
            if not date_val:
                continue

            existing = (
                db.query(Activity)
                .filter(
                    Activity.task_id == task_id,
                    Activity.date == date_val,
                    Activity.start_time == start_time_cell,
                )
                .first()
            )

            if not existing:
                activity = Activity(
                    task_id=task_id,
                    date=date_val,
                    start_time=start_time_cell,
                    end_time=end_time_cell,
                    comments=comments_cell,
                    links=links_cell,
                )
                db.add(activity)
                activity_count += 1
                file_activity_count += 1

                if activity_count % 100 == 0:
                    db.commit()

        print(f"  ✓ Loaded {file_activity_count} activities")

    db.commit()
    print(f"✓ Total: {activity_count} activities loaded")

    db.close()
    print("✓ Initial data load complete!\n")


def init_db():
    """Initialize database - creates tables and loads Excel data on first run only"""
    Base.metadata.create_all(bind=get_engine())

    db = get_session()

    # Check if data has already been loaded (using a persistent flag)
    data_loaded_flag = db.query(Settings).filter(Settings.key == "excel_data_loaded").first()

    if not data_loaded_flag:
        # First run - load data from Excel files
        print("🔄 Loading data from Excel files (one-time)...\n")
        load_excel_data()

        # Mark data as loaded so it never loads again
        loaded_flag = Settings(key="excel_data_loaded", value="true")
        db.add(loaded_flag)
        db.commit()
        print("✅ Data loaded successfully. This will not happen again.\n")
    else:
        # Data already loaded - just skip to using the database
        task_count = db.query(Task).count()
        activity_count = db.query(Activity).count()
        print(f"✅ Using existing database ({task_count} tasks, {activity_count} activities)\n")

    db.close()


if __name__ == "__main__":
    init_db()
